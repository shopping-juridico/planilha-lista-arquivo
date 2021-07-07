from openpyxl import load_workbook

wb = load_workbook('excel files/lista-arquivo-teste.xlsx')

ws = wb.active

for cell in ws["A"]:
    a = cell.value
    fatia_a = a[len(a) - 5:len(a)] #localiza a parte a ser identificada
    if ")" in fatia_a:
        b = cell.value
        fatia_b = b[len(b) - 15:len(b) - 4] #identifica a parte a ser removida
        c = cell.value
        c = c.replace(fatia_b, '') #remove a parte identificada
        cell.value = c
        print(c)
        
wb.save('excel files/lista-arquivo-novo.xlsx')
